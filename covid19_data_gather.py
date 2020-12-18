############
## IMPORTS ####################################################################
# data i/o
from pathlib import Path
import sys
import csv
import copy
import os.path
import shutil

# date math
import datetime
from datetime import date, datetime, timedelta

# json
import json
import fastjsonschema

# NYT git clone
from git import Repo

# xlsx handling
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
    Series,
)

# Email send c/o outlook
import win32com.client
from win32com.client import Dispatch, constants

# Enumerations
from enum import Enum

#########
## todo #######################################################################
#########

############
## GLOBALS ####################################################################
# Path Settings
g_py_path = Path(__file__).resolve().parent
g_json_path = g_py_path / "covid19_data_gather_conf.json"
g_json_path_orig = g_json_path
g_json_path_default = g_py_path / "sample_covid19_data_gather_conf.json"
g_json_schema_path = g_py_path / "covid19_data_gather_conf.schema.json"
g_xlsx_path = g_py_path / "xlsx"
g_root_path = g_py_path.parent
g_covid19_data_path = g_root_path / "covid-19-data"

# Global App Settings
# g_debug = True
g_debug = False
g_date_fmt = '%Y-%m-%d'  # As per NYT data (YYYY-MM-DD)

# Set to false to disable json validation should it prove troublesome in the future.
g_json_validate = True


## JSON Based Conf ###############################################################
# App settings should be stored in covid19_data_gather_conf.json in a "settings": { ... } block.
#
# Each key in "settings" { "key": "value "} should have a global variable defined below, set to a pre-json load default.
#
# The default below is used in the event the settings key is missing from the configuration file.
#
# When adding a new g_conf["settings"][<full key name>]:
#   1. Initialize g_<abbr key name> below
#   2. To assign_global_vars(), add:
#      * global g_<abbr key name>
#      * g_<abbr key name> = get_global_conf(SETTING, '<full key name>', g_<abbr key name>)
#   3. Use g_<abbr key name> as normal in rest of program.

g_conf = dict()

# Default Settings
# Minimum case count to be charted; set to 1 for accurate active counts
g_case_benchmark = 1
g_days = 28  # 4 weeks * 7 days per week
g_per_county = 100000
g_per_state = 100000

g_email = False
g_email_client = "N/A"
g_email_to = []
g_email_style = "font-family: Trebuchet MS; color:#25253b; font-size:14pt"
g_email_greeting = "Hello!<br>"
g_email_sig = "xoxo<br>Yours Truly"


# Metadata
g_state_abbr = ["AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS",
                "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY"]
g_state_name = {"AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas", "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware", "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland", "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi", "MO": "Missouri", "MT": "Montana",
                "NE": "Nebraska", "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina", "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah", "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia", "WI": "Wisconsin", "WY": "Wyoming"}

#########
# Debug #######################################################################


def dbg(s):
    if g_debug:
        print(s)

##########################
# Logging & Benchmarking ######################################################
#
# br()
# t=log_start('doing something')
# ...
# log_end(t)
#


def abort():
    print()
    print("Aborting.")
    print()
    exit(1)


def br():
    print()


def log_start(s):
    t_start = datetime.now()
    print(s, end='', flush=True)
    return t_start


def log_end(t_start):
    t_end = datetime.now()
    delta = t_end-t_start
    print(' ['+str(delta)+']')

###############
# json config #################################################################

# settings - main configuraation block
# anything else - spreadsheet generator
#


def validate_json():
    if not g_json_validate:
        return

    global g_conf

    # load json conf schema into memory
    s = ""
    with open(g_json_schema_path) as f:
        s = json.load(f)
    f.close()

    # validate json conf vs schema
    v = fastjsonschema.compile(s)
    try:
        v(g_conf)
    except fastjsonschema.JsonSchemaException as e:
        print()
        print()
        print(f"Error - Invalid {g_json_path}")
        print(f"        -> \"{e.name}\": \"{e.value}\" <-")
        print(f"        {e.message}")
        print()
        print(f"Please update {g_json_path} and try again!")
        abort()


def get_global_conf(conf_block, setting, default):
    try:
        return g_conf[conf_block][setting]
    except Exception:
        return default


def assign_global_vars():
    # Assign global settings from json
    global g_conf
    global g_email
    global g_email_client
    global g_email_to
    global g_email_style
    global g_email_greeting
    global g_email_sig
    global g_case_benchmark
    global g_days
    global g_per_county
    global g_per_state

    SETTING = 'settings'

    g_email = get_global_conf(SETTING, 'send-email', g_email_to)
    g_email_client = get_global_conf(SETTING, 'send-email-client', g_email_to)
    g_email_to = get_global_conf(SETTING, 'send-email-to', g_email_to)  # array
    g_email_style = get_global_conf(SETTING, 'send-email-style', g_email_style)
    g_email_greeting = get_global_conf(
        SETTING, 'send-email-greeting', g_email_greeting)
    g_email_sig = get_global_conf(SETTING, 'send-email-signature', g_email_sig)
    g_case_benchmark = get_global_conf(
        SETTING, 'case-min-benchmark', g_case_benchmark)
    g_days = get_global_conf(SETTING, 'case-days-duration', g_days)
    g_per_county = get_global_conf(
        SETTING, 'geography-per-county', g_per_county)
    g_per_state = get_global_conf(SETTING, 'geography-per-state', g_per_state)


def load_configuration():
    global g_conf
    global g_json_path

    cmdline_override = False

    if len(sys.argv) > 1:
        potential_json_fn = sys.argv[1]
        if not os.path.exists(potential_json_fn):
            print()
            print(f"{potential_json_fn} does not exist.")
            print(f"Falling back to {g_json_path}")
            print()
        else:
            g_json_path = sys.argv[1]
            cmdline_override = True
            print()
            print(f"Using config override: {g_json_path}")
            print()

    if not cmdline_override and not os.path.exists(g_json_path):
        if os.path.exists(g_json_path_default):
            print()
            t_copy = log_start(
                f"No custom configuration found at {g_json_path} -- create")
            shutil.copy(g_json_path_default, g_json_path)
            log_end(t_copy)
            print()
            print(f"Be sure to modify {g_json_path} to suit your needs!")
            print()
        else:
            print()
            print(
                f"No custom configuration found at {g_json_path}, nor is there a default file at {g_json_path_default}.")
            print()
            print("Consider recloning from https://github.com/const-void/covid-19-spreadsheet or supplying a configuration file via command line.")
            print()
            abort()

    t = log_start('Load settings')

    with open(g_json_path) as f:
        g_conf = json.load(f)
    f.close()

    validate_json()

    assign_global_vars()

    log_end(t)


#######
# git #########################################################################

# git pull remote for nyt data


def update_data():
    br()
    t = log_start("Checking nyt source data... ")
    r = Repo(g_covid19_data_path)
    cur_head = r.head.commit
    fetches = r.remotes.origin.pull()

    if cur_head == r.head.commit:
        print("No updates.")
    else:
        for f in fetches:
            print("Updated %s to %s - %s" % (f.ref, f.commit, f.note))

    log_start("Complete!")
    log_end(t)

#
# Todo - update getters/setters to properties

##################
### Date Caching ##############################################################

# Active = YTD (Today) - YTD (28 days ago)
# To simplify the math we map 'today' to '28 days in the future'.  Ok, this is probably complicating it but???
# Since we are processing thousands of cases, which potentially could be on the same day,
# we don't want to perform this expensive date calculation with each record.
# Instead, we cache Today => 28 days into the future in a global dictionary for use by the county class.
# states are the aggregation of counties, and don't need to do this calculation at the state level.


g_active_date_map = dict()


def cache_active_to_inactive_date_map():
    br()
    t = log_start('Caching active to inactive dates')

    global g_active_date_map
    g_active_date_map = dict()

    # From Jan-2020 to Today
    date_active = date(2020, 1, 1)
    # Assume consant illness window of g_days
    date_inactive = date_active+timedelta(days=g_days)
    date_now = date.today()

    # Increment a day at a time
    date_delta = timedelta(days=1)

    date_active_str = ""
    date_inactive_str = ""

    while date_active <= date_now:
        # str conversion
        date_active_str = date_active.strftime(g_date_fmt)
        date_inactive_str = date_inactive.strftime(g_date_fmt)

        # map
        # print(date_active_str+" => "+date_inactive_str)
        g_active_date_map[date_active_str] = date_inactive_str

        # increment
        date_active += date_delta
        date_inactive += date_delta

    log_end(t)

######################
## Geography Classes ##########################################################


class State:
    region = 0
    division = 0
    state_fips = 0
    population = 0
    name = ""
    counties = []
    covid19_cases = dict()
    cur_covid19_case = None
    prior_covid19_case = None
    geography = "State"
    unknown_county = None

    def set(self, region, division, state_fips, name):
        self.region = int(region)
        self.division = int(division)
        self.state_fips = int(state_fips)
        self.name = name
        self.counties = []
        self.covid19_cases = dict()
        self.population = 0

    def __init__(self, data_array):
        # print(data_array)
        self.set(data_array[0], data_array[1], data_array[2], data_array[3])

    # Accessors

    # used by County -- todo, remove
    def get(self, mips):
        return self

    def get_counties(self):
        return self.counties

    def get_all_counties(self):
        c = []
        for county in sorted(self.counties, key=lambda x: x.name):
            if county.get_population() > 0 and county.has_covid19():
                c.append(county)
        return c

    def get_geography(self):
        return self.geography

    def get_fips(self):
        return self.state_fips

    def get_name(self):
        return self.name

    def get_population(self):
        return self.population

    def get_parent_location(self):
        return "US"

    def get_prior_covid19_case(self):
        if self.prior_covid19_case is not None:
            return self.prior_covid19_case
        else:
            return zero_covid19

    def get_specific_covid19_case(self, d):
        if d in self.covid19_cases.keys():
            return self.covid19_cases[d]
        else:
            return zero_covid19

    def get_current_covid19_case(self):
        return self.cur_covid19_case

    # Setters
    def set_unknown_county(self, c):
        self.unknown_county = c
        self.add_county(c)

    def add_county(self, c):
        self.counties.append(c)
        # if self.get_name()=="Florida":
        #   print("Adding %s to %s (%i)"%(c.get_name(),self.get_name(),len(self.counties)))

    def add_covid19_case(self, covid19_case):
        if covid19_case.case_date not in self.covid19_cases:
            self.covid19_cases[covid19_case.case_date] = copy.copy(
                covid19_case)
        else:
            self.covid19_cases[covid19_case.case_date].add(covid19_case)

        # todo - should add data logic to make sure data is not out of order

        # if we have a new day of covid19 data, 'current' day is really the prior day
        if self.cur_covid19_case is not None:
            if self.cur_covid19_case.case_date != covid19_case.case_date:
                self.prior_covid19_case = self.cur_covid19_case

        self.cur_covid19_case = self.covid19_cases[covid19_case.case_date]

    def set_population(self, population):
        self.population = population
        self.unknown_county.set_population(
            population)  # technically..true, right?

    def has_covid19(self):
        return len(self.covid19_cases) > 0


class States:
    input = "state-geocodes-v2018.csv"
    pop_input = "nst-est2019-01.csv"
    states = dict()
    states_by_name = dict()

    def load_states(self):
        with open(self.input) as csv_file:
            csv_data = csv.reader(csv_file)
            for r in csv_data:
                s = State(r)
                self.states[s.state_fips] = s
                self.states_by_name[s.name] = s

                # unknown county setup
                u = copy.copy(unknown_county_data)
                u[1] = s.state_fips
                s.set_unknown_county(County(u, self))

    def load_population(self):
        with open(self.pop_input) as csv_file:
            csv_file.readline()
            csv_data = csv.reader(csv_file)
            for r in csv_data:
                state_name = r[0]
                state_pop = int(r[12])
                state = self.get_by_name(state_name)
                if state is None:
                    dbg("state population load - couldn't find %s (pop: %d), skipping" %
                        (state_name, state_pop))
                else:
                    state.set_population(state_pop)

    def __init__(self):
        self.states = dict()
        self.states_by_name = dict()
        self.load_states()
        self.load_population()

    def get(self, state_fips):
        if state_fips not in self.states:
            return None
        return self.states[state_fips]

    def get_by_name(self, n):
        if n not in self.states_by_name.keys():
            return None
        return self.states_by_name[n]

    def get_all_states(self):
        all_states = []
        for state_name in sorted(self.states_by_name):
            state = self.states_by_name[state_name]
            if state.has_covid19():
                all_states.append(state)

        return all_states


unknown_county_data = [0, 0, 0, 0, 0, 0, 'Unknown County']


class County:
    summary_level = 0
    state_fips = 0
    county_fips = 0
    county_subdivision_fips = 0
    place_fips = 0
    conslidated_city_fips = 0
    name = ''
    state = None
    population = 0
    covid19_cases = []
    covid19_cases_inactive = dict()
    prior_covid19_case = None
    cur_covid19_case = None
    geography = "County"

    def set(self, summary_level, state_fips, county_fips, county_subdivision_fips, place_fips, conslidated_city_fips, area_name):
        self.covid19_cases = []
        self.covid19_cases_inactive = dict()
        self.covid19_cases_by_date = dict()

        self.summary_level = int(summary_level)
        self.state_fips = int(state_fips)
        self.county_fips = int(county_fips)
        self.county_subdivision_fips = int(county_subdivision_fips)
        self.place_fips = int(place_fips)
        self.conslidated_city_fips = int(conslidated_city_fips)
        self.name = area_name

        # if self.place_fips==0 and self.county_fips!=0:
        self.place_fips = int(state_fips+county_fips)

    def __init__(self, data_row, states):
        self.set(data_row[0], data_row[1], data_row[2],
                 data_row[3], data_row[4], data_row[5], data_row[6])

        # link state and counties together
        self.state = states.get(self.state_fips)

    def get_state(self):
        return self.state

    def get_geography(self):
        return self.geography

    def set_population(self, estimated_population):
        self.population = int(estimated_population)

    def add_covid19_case(self, covid19_case):
        self.prior_covid19_case = self.cur_covid19_case

        # inactive covid19 handling
        idx = g_active_date_map[covid19_case.case_date]
        self.covid19_cases_inactive[idx] = copy.copy(covid19_case)

        if covid19_case.case_date in self.covid19_cases_inactive:
            covid19_case.active_count = covid19_case.case_count - \
                self.covid19_cases_inactive[covid19_case.case_date].case_count
        else:
            covid19_case.active_count = covid19_case.case_count

        self.cur_covid19_case = copy.copy(covid19_case)
        self.covid19_cases.append(covid19_case)
        self.covid19_cases_by_date[covid19_case.case_date] = copy.copy(
            covid19_case)

        if self.state is not None:
            self.state.add_covid19_case(covid19_case)
        else:
            dbg("County %s has unknown state" % (self.name))

    def has_covid19(self):
        return len(self.covid19_cases) > 0

    def get_prior_covid19_case(self):
        if self.prior_covid19_case is None:
            return zero_covid19
        else:
            return self.prior_covid19_case

    def get_current_covid19_case(self):
        return self.cur_covid19_case

    def get_specific_covid19_case(self, d):
        if d in self.covid19_cases_by_date.keys():
            return self.covid19_cases_by_date[d]
        else:
            return zero_covid19

    def get_name(self):
        return self.name

    def get_parent_location(self):
        return self.state.get_name()

    def get_location(self):
        return "%s, %s" % (self.get_name(), self.state.get_name())

    def get_population(self):
        return self.population

    def get_fips(self):
        return self.place_fips

    def get_csv_output(self):
        return "%s,%i,%i" % (self.get_location(), self.get_fips(), self.get_population())

    def get_covid19_cases(self):
        return self.covid19_cases


class Counties:
    input = "all-geocodes-v2018.csv"
    counties_by_fips = dict()
    counties_for_pop_est = dict()

    def __init__(self, states):
        self.counties_by_fips = dict()
        self.counties_for_pop_est = dict()
        self.counties_by_name = dict()
        with open(self.input) as csv_file:
            csv_data = csv.reader(csv_file)
            for r in csv_data:
                # 010,00,000,00000,00000,00000,United States
                # 040,01,000,00000,00000,00000,Alabama
                # print(r)
                c = County(r, states)
                if c.state is not None:
                    c.state.add_county(c)
                    # if c.state.get_name()=="Florida":
                    # print(c.state,len(c.state.counties),c.state.get_name())

                # for nyt
                # first place_fips wins, I guess?
                if c.place_fips not in self.counties_by_fips:
                    self.counties_by_fips[c.place_fips] = c

                # for pop
                s = c.get_state()
                if s is not None:
                    state_fips = s.get_fips()
                    if state_fips not in self.counties_for_pop_est:
                        self.counties_for_pop_est[state_fips] = dict()

                    self.counties_for_pop_est[state_fips][c.name] = c

    def get(self, place_fips):
        fips = int(place_fips)
        if fips not in self.counties_by_fips.keys():
            return None
        return self.counties_by_fips[place_fips]

    def get_for_pop_est(self, state_fips, county_name):
        if state_fips not in self.counties_for_pop_est:
            return None
        if county_name not in self.counties_for_pop_est[state_fips]:
            return None

        return self.counties_for_pop_est[state_fips][county_name]


def set_county_population(states, counties):
    input = "co-est2019-annres.csv"
    with open(input) as csv_file:
        csv_data = csv.reader(csv_file)
        for r in csv_data:
            location_name = r[0]
            population = int(float(r[12]))
            if "," in location_name:
                (county_name, state_name) = location_name.split(",")
                county_name = county_name.replace(".", "").strip()
                state_name = state_name.strip()
                state = states.get_by_name(state_name)
                if state is None:
                    dbg("county population - Unknown state %s, skipping %s, %s" %
                        (state_name, state_name, county_name))
                else:
                    county = counties.get_for_pop_est(
                        state.get_fips(), county_name)
                    if county is None:
                        dbg("county population - Unknown county %s, skipping %s, %s" %
                            (county_name, state_name, county_name))
                    else:
                        # print("%s, %s - %i "%(state.get_name(),county_name,population))
                        county.set_population(population)


def validate_custom_geographies(states, counties):
    if not g_json_validate:
        return

    global g_conf

    has_err = False
    for custom_xlsx in g_conf['spreadsheets']['custom']:
        for geography in g_conf['spreadsheets']['custom'][custom_xlsx]:
            # validate counties - json schema handles states
            if len(geography) != 2:
                # we have a county in <County>, <State Abbr> form  - json schema validation gets us to this point.

                # Split geography into component parts
                csv = geography.split(",")
                county = csv[0].strip()
                state_abbr = csv[1].strip()

                # Validate state
                if state_abbr not in g_state_abbr:
                    # Bad state
                    print()
                    print()
                    print(f"Error - Invalid state in {g_json_path}")
                    print(
                        f"        -> spreadsheets.custom.{custom_xlsx} = [ ... \"{geography}\" ... ] <-")
                    print(
                        f"        {geography} has an unknown state abbreviation [{state_abbr}]")
                    print(
                        f"        State abbreviation must be one of: {g_state_abbr}")
                    print()
                    has_err = True
                else:
                    # Validate county
                    state_fips = states.states_by_name[g_state_name[state_abbr]].state_fips
                    county_list = counties.counties_for_pop_est[state_fips]
                    if county not in county_list:
                        # Bad county
                        # Filter valid County geographies to those that include the string "County"
                        valid_counties = sorted(
                            [county for county in county_list.keys() if "County" in county])
                        print()
                        print()
                        print(f"Error - Invalid county in {g_json_path}")
                        print(
                            f"        -> spreadsheets.custom.{custom_xlsx} = [ ... \"{geography}\" ... ] <-")
                        print(
                            f"        {geography} has an unknown county [{county}]")
                        print(f"      {state_abbr} counties must be one of:")
                        print(f"{valid_counties}")
                        print()
                        has_err = True

    if has_err:
        print(f"Please update {g_json_path} and try again!")
        abort()

##########################
## Covid Case Statistics ######################################################


class Covid19Stat:
    case_date = ''
    case_count = 0
    death_count = 0
    active_count = 0

    def __init__(self, case_date, case_count, death_count):
        # self.date=datetime.datetime.strptime(case_date,"%Y-%m-%d")
        if (death_count == ''):
            death_count = 0
        if (case_count == ''):
            case_count = 0

        self.case_date = case_date
        self.case_count = int(case_count)
        self.death_count = int(death_count)
        self.active_count = 0

    def get_csv_output(self):
        return "%s,%i,%i" % (self.case_date, self.case_count, self.death_count)

    def add(self, covid19_case):
        self.case_count += covid19_case.case_count
        self.death_count += covid19_case.death_count
        self.active_count += covid19_case.active_count


# Used for the prior day of the first day of a case.
zero_covid19 = Covid19Stat('1900-01-01', 0, 0)


def set_county_covid19_cases(states, counties):
    input = g_covid19_data_path / "us-counties.csv"
    with open(input) as csv_file:
        csv_file.readline()  # skip first line
        csv_data = csv.reader(csv_file)
        for r in csv_data:
            #date,        county,     state,       fips,   cases,deaths
            # 2020-01-21,  Snohomish,  Washington,  53061,  1,    0
            #r[0],      r[1],       r[2],        r[3],   r[4], r[5]

            # try:
            covid_case = Covid19Stat(r[0], r[4], r[5])
            # except ValueError:
            #    print('error!')
            #    print(r)
            #    raise

            covid_county = r[1].strip()
            covid_state = r[2].strip()
            covid_fips = r[3]
            if len(covid_fips) != 0:
                covid_fips = int(covid_fips)
                county = counties.get(covid_fips)
                if county is None:
                    # unknown county
                    dbg("covid19 case data - unknown county %s, %s (%i)" %
                        (covid_county, covid_state, covid_fips))
                    state = states.get_by_name(covid_state)
                    if state is None:
                        # unknown state
                        dbg("covid19 case data - unknown county & state %s, %s (%i)" %
                            (covid_county, covid_state, covid_fips))
                    else:
                        state.unknown_county.add_covid19_case(covid_case)
                else:
                    county.add_covid19_case(covid_case)

                    # county trap
                    # if covid_county == 'Burleigh':
                    #    print(
                    #        f"Burleigh = {r}, {county.name} {county.state.name} {county.place_fips}")
            else:
                # covid_county is 'Unknown' or actually uknown
                state = states.get_by_name(covid_state)
                if state is not None:
                    state.unknown_county.add_covid19_case(covid_case)
                else:
                    dbg("covid19 case data - missing fips and uknown state %s, %s (%s)" %
                        (covid_county, covid_state, covid_fips))

################
## Spreadsheet ################################################################


class XLSX:

    def init_data(self):
        # data dictionaries have a key of YYYY-MM-DD and an integer value representing covid19 data.
        self.death_per_capita = dict()
        self.cfr = dict()
        self.per_capita_cases = dict()

        self.actual_cases_new = dict()  # Increase/decrease from prior day (daily delta)
        self.death_cases_new = dict()  # increase/decrease from prior day (daily delta)

        self.active_cases = dict()  # YTD
        self.actual_cases = dict()  # YTD
        self.death_cases = dict()  # YTD

        # Must sequence/order of self.data_wb
        self.data = [self.death_per_capita, self.cfr, self.per_capita_cases,
                     self.actual_cases_new, self.death_cases_new,
                     self.active_cases, self.actual_cases, self.death_cases]

        self.per_capita_by_geography = dict()
        self.per_capita_by_geography['State'] = g_per_state
        self.per_capita_by_geography['County'] = g_per_county
        self.per_capita_str = self.humanize(
            self.per_capita_by_geography['County'])

    def init_wb(self):
        self.wb = Workbook()

    def init_ws(self):
        # summary worksheet
        self.cur_data_wb = self.wb.active
        self.cur_data_wb.title = "today"

        # todo - create an object representing data dictionary, workbook and chartname
        # data worksheets
        self.active_wb = self.wb.create_sheet(title="active")
        self.death_per_capita_wb = self.wb.create_sheet(
            title="death per %s" % (self.per_capita_str))
        self.actual_new_wb = self.wb.create_sheet(title="actual delta")
        self.cfr_wb = self.wb.create_sheet(title="case fatality rate")
        self.per_capita_wb = self.wb.create_sheet(
            title="reported per %s" % (self.per_capita_str))
        self.death_new_wb = self.wb.create_sheet(title="dead delta")
        self.death_wb = self.wb.create_sheet(title="dead")
        self.actual_wb = self.wb.create_sheet(title="actual")

        # must match sequence/order of self.data
        self.data_wb = [self.death_per_capita_wb, self.cfr_wb, self.per_capita_wb,
                        self.actual_new_wb, self.death_new_wb,
                        self.active_wb, self.actual_wb, self.death_wb]

        # chart name sequence must match sequence/order of self.data_wb
        self.chart_names = ["Death Per %s" % (self.per_capita_str), "Case Fatality Rate", "Reported Per %s" % (self.per_capita_str),
                            "Daily Reported", "Daily Dead",
                            "Active", "Reported", "Dead"]

        self.changelog_wb = self.wb.create_sheet(title="changelog")

    def gen_header_row(self):
        self.hdr = ["Day"]
        self.loc_names = []
        for loc in self.locations:
            self.hdr.append(loc.get_name())
            self.loc_names.append(loc.get_name())

    def add_headers(self):
        self.gen_header_row()
        for wb in self.data_wb:
            wb.append(self.hdr)

    # Build an empty day for each geography in each data dictionary
    def check_day(self, day):
        for case_dict in self.data:
            if day not in case_dict:
                case_dict[day] = dict()
                for n in self.loc_names:
                    case_dict[day][n] = None  # 0

    def gen_location_data(self, location):
        # init
        cur_case = 0
        loc_name = location.get_name()
        loc_per_capita = self.per_capita_by_geography[location.get_geography(
        )]/location.get_population()
        prior_day_case = None

        # calculate covid19 case data for geography cases
        for covid_data in location.covid19_cases:
            covid_case = covid_data
            if isinstance(covid_data, str):
                covid_case = location.covid19_cases[covid_data]

            if covid_case.case_count >= self.benchmark:
                cur_case += 1
                cur_day = "Day %03d" % (cur_case)

                self.check_day(cur_day)

                # value calcs
                self.death_per_capita[cur_day][loc_name] = covid_case.death_count*loc_per_capita
                self.cfr[cur_day][loc_name] = covid_case.death_count / \
                    covid_case.case_count*100
                if location.get_population() > 0:
                    self.per_capita_cases[cur_day][loc_name] = covid_case.case_count*loc_per_capita
                self.active_cases[cur_day][loc_name] = covid_case.active_count
                self.actual_cases[cur_day][loc_name] = covid_case.case_count
                self.death_cases[cur_day][loc_name] = covid_case.death_count

                # new cases = todays YTD count - yesterday's YTD count
                if prior_day_case is not None:
                    self.actual_cases_new[cur_day][loc_name] = covid_case.case_count - \
                        prior_day_case.case_count
                    self.death_cases_new[cur_day][loc_name] = covid_case.death_count - \
                        prior_day_case.death_count

                prior_day_case = covid_case

    def gen_locations_data(self):
        for loc in self.locations:
            self.gen_location_data(loc)

    def add_count_to_xlsx(self, ws, data, title):
        for day in sorted(data):
            r = [day]
            for n in self.loc_names:
                r.append(data[day][n])
            ws.append(r)

        # Build Chart
        chart = LineChart()
        chart.title = "%s Covid-19 Cases" % (title)
        chart.style = 13
        chart.y_axis.title = "%s Count" % (title)
        chart.x_axis.title = "Days"
        cats = Reference(ws, min_col=1, min_row=1,
                         max_col=1, max_row=len(data)+1)
        data = Reference(ws, min_col=2, min_row=1,
                         max_col=len(self.hdr), max_row=len(data)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "G2")

    def add_counts_to_xlsx(self):
        for idx in range(len(self.data)):
            self.add_count_to_xlsx(
                self.data_wb[idx], self.data[idx], self.chart_names[idx])

    def add_population_to_xlsx(self):
        col_num = Enum(
            'COL_NUM', ['NAME',
                        'REGION',
                        'CASE_DATE',
                        'DEAD',
                        'NEW_DEAD',
                        'DEAD_PER',
                        'CFR',
                        'CASES',
                        'ACTIVE_CASES',
                        'AVG_DLY_RATE',
                        'NEW_CASES',
                        'ACTIVE_PER',
                        'CASES_PER',
                        'POP',
                        'GEO_PER',
                        'PER_SCALE'])

        self.cur_data_wb.append(['Name',                                            # 0
                                 'Containing Region',                               # 1
                                 'Case Date',                                       # 2
                                 'Dead',                                            # 3
                                 'New Dead',                                        # 4
                                 'Dead Per %s' % (
                                     self.per_capita_str),                          # 5
                                 'Case Fatality Rate',                              # 6
                                 'Cases',                                           # 7
                                 'Active Cases',                                    # 8
                                 'Daily Avg',                                       # 9 - 10/10/2020
                                 'New Cases',                                       # 10
                                 'Trend',                                           # 11 - 10/10/2020
                                 'Active Per %s' % (
                                     self.per_capita_str),                          # 12
                                 'Cases Per %s' % (
                                     self.per_capita_str),                          # 13
                                 'Population',                                      # 14
                                 'Geography Per %s Scale' % (
                                     self.per_capita_str),                          # 15
                                 'Per %s' % (self.per_capita_str)])                 # 16

        # To do - automatically enumerate below
        # Manually update below column 'constants' after updating above
        COL_NEW_DEAD_CHAR = 'D'
        COL_CASES_CHAR = 'H'

        # Col Format Params
        self.sorting_col_num = col_num.AVG_DLY_RATE.value-1
        self.percent_col_num = col_num.CFR.value
        self.neg_col_nums = (col_num.AVG_DLY_RATE.value,)
        self.comma_col_nums = (col_num.DEAD.value,
                               col_num.NEW_DEAD.value,
                               col_num.CASES.value,
                               col_num.ACTIVE_CASES.value,
                               col_num.NEW_CASES.value,
                               col_num.ACTIVE_PER.value,
                               col_num.CASES_PER.value)
        self.scale_cols = (COL_NEW_DEAD_CHAR,
                           COL_CASES_CHAR)

        cell_data = []
        for loc in self.locations:
            # loc=self.locations[loc_name]
            g = loc.get_geography()
            p = loc.get_population()  # p (population)
            per_capita_scale = self.per_capita_by_geography[g]
            per_capita = round(p / per_capita_scale, 2)
            d = loc.get_current_covid19_case()  # d (data)
            if d is None:
                print("Got None")
                print(loc.name)
                print(len(loc.covid19_cases))
            prior_d = loc.get_prior_covid19_case()  # prior_d (prior_data)
            if d.case_count != 0:
                d_ratio = d.death_count/d.case_count
            else:
                if d.death_count == 0:
                    d_ratio = 0.0
                else:
                    d_ratio = 1.0

            # Get last week's date
            case_date_obj_cur = datetime.strptime(d.case_date, g_date_fmt)
            case_date_obj_last_week = case_date_obj_cur - timedelta(days=7)
            case_date_last_week = case_date_obj_last_week.strftime(g_date_fmt)

            case_count_new = d.case_count - prior_d.case_count

            # Calculate 7 day activity average
            last_week = loc.get_specific_covid19_case(case_date_last_week)
            wkly_avg = (d.active_count - last_week.active_count)/7

            # Derived measures
            net_zero = case_count_new - wkly_avg
            if net_zero != 0:
                sick_percentage = wkly_avg / net_zero
            elif wkly_avg != 0:
                # 100 new cases, 100 wkly avg: 100% sick
                sick_percentage = 1.0
            else:
                # 0 new cases, 0 wkly avg: 0% sick.
                sick_percentage = 0.0
            sick_ratio = int(sick_percentage)

            # Identify trend based on weekly avg
            trend = ""
            if wkly_avg > -1:  # a positive avg means more infections! Label/
                if sick_percentage >= 0.9:
                    trend = "UNCONTROLLED"
                elif sick_percentage >= 0.50:
                    trend = "DANGER ZONE"
                elif sick_percentage >= 0.25:
                    trend = "ACTIVE SPREAD"
                elif sick_percentage >= 0.1:
                    trend = "WARNING"
                else:
                    trend = "CONTROLLED"

                # if we have a daily avg of 200, and a new case count, then today is a test anamoly.
                if wkly_avg >= case_count_new:
                    if wkly_avg > 1000:
                        trend = "UNCONTROLLED"
                    elif wkly_avg > 500:
                        trend = "DANGER ZONE"
                    elif wkly_avg > 100:
                        trend = "ACTIVE SPREAD"
                    elif wkly_avg > 50:
                        trend = "WARNING"
                    elif wkly_avg > 10:
                        trend = "TRYING"
                    else:
                        trend = "CONTROLLED"

                # If case_count = 100, and 25 people are daily avg, then 75 people who got sick replace people just getting healthy.
                # this 75 is the "net_zero" -- the virus is not going up nor down, but staying the same.

                trend_summary = ""
                if net_zero > 0:
                    # if case_count = 100, daily_avg = 50:
                    #   50 people got better
                    #   50 people got sick
                    #   50 more people got sick
                    #  then sick to healthy ratio is 1:1
                    # if case_count = 100, daily_avg = 75, then:
                    #   25 people got better
                    #   25 people got sick
                    #   75 more people got sick
                    #  for each net_zero, 3 more people got sick

                    if sick_ratio >= 1:
                        sick_ratio = sick_ratio+1
                        trend_summary = f"{sick_ratio}:1 growth"
                    else:
                        trend_summary = "+{0:.0%} growth".format(
                            sick_percentage)
                else:
                    trend_summary = "n/a"

                trend = f"{trend} [{trend_summary}]"

            else:
                cure_days = int(d.active_count/abs(wkly_avg))
                cure_date = (case_date_obj_cur +
                             timedelta(days=cure_days)).strftime(g_date_fmt)
                trend = f"{cure_days} days ({cure_date})"

            cell_data.append([loc.get_name(),  # 0
                              loc.get_parent_location(),  # 1
                              d.case_date,  # 2
                              d.death_count,  # 3
                              d.death_count - prior_d.death_count,  # 4
                              round(d.death_count/per_capita, 2),  # 5
                              d_ratio,  # 6
                              d.case_count,  # 7
                              d.active_count,  # 8
                              wkly_avg,  # 9 - 10/10/2020
                              case_count_new,  # 10
                              trend,  # 11 - 10/10/2020
                              round(d.active_count/per_capita, 2),  # 12
                              round(d.case_count/per_capita, 2),  # 13
                              self.humanize(p),  # 14
                              self.humanize(per_capita_scale),  # 15
                              per_capita])  # 16

        cell_data.sort(key=lambda x: x[self.sorting_col_num], reverse=True)
        for data in cell_data:
            self.cur_data_wb.append(data)

        self.autosize_wb_cols(self.cur_data_wb)
        self.filter_wb_cols(self.cur_data_wb)

    def humanize(self, n):
        thousands = 1000
        millions = 1000000
        if n < thousands:
            return str(n)

        if n < millions:
            units = "k"
            val = f"{n/thousands:.1f}"
        else:
            units = "M"
            val = f"{n/millions:,.1f}"

        return "%s%s" % (val.rstrip('0').rstrip('.'), units)

    def autosize_wb_cols(self, wb):
        fills = [openpyxl.styles.PatternFill("solid", "548235"),   # -100 or better
                 openpyxl.styles.PatternFill("solid", "A9D08E"),   # -40...-99
                 openpyxl.styles.PatternFill("solid", "C6E0B4"),   # -5...-39
                 openpyxl.styles.PatternFill("solid", "FFD966"),   # 4...-5
                 openpyxl.styles.PatternFill("solid", "FFCCCC"),   #
                 openpyxl.styles.PatternFill("solid", "FF7C80"),   #
                 openpyxl.styles.PatternFill("solid", "FF0000")
                 ]

        for column_cells in wb.columns:
            max_len = 0
            col_letter = column_cells[0].column_letter

            for cell in column_cells:
                cur_len = len(str(cell.value))
                if cur_len > max_len:
                    max_len = cur_len

                # Cell Styling
                # number_format=https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
                cur_col = cell.column
                if cur_col in self.comma_col_nums:
                    cell.style = 'Comma [0]'
                elif cur_col in self.neg_col_nums:
                    cell.number_format = '#,##0'
                    # Determine bg color
                    try:
                        val = int(cell.value)
                        if val < -100:
                            cell.fill = fills[0]
                        elif val < -40:
                            cell.fill = fills[1]
                        elif val < - 5:
                            cell.fill = fills[2]
                        elif val < 5:
                            cell.fill = fills[3]
                        elif val < 40:
                            cell.fill = fills[4]
                        elif val < 100:
                            cell.fill = fills[5]
                        else:
                            cell.fill = fills[6]
                    except:
                        pass  # ignore

                elif cur_col == self.percent_col_num:
                    cell.number_format = '0.00%'

            # Need to account for columns with commas
            scale_factor = 1
            if col_letter in self.scale_cols:
                scale_factor = 1.4
            wb.column_dimensions[col_letter].width = (max_len+2)*scale_factor

    def filter_wb_cols(self, wb):
        wb.auto_filter.ref = wb.dimensions

    def add_change(self, date, desc):
        self.changelog_wb.append([date, desc])

    def add_changelog_to_xlsx(self):
        # header
        self.add_change('Date', 'Description')

        # Changes
        self.add_change(
            "2020-12-17", "covid19_data_gather_conf.json: Implemented full validation via schema when possible, and geography data when not.")
        self.add_change(
            "2020-12-15", "covid19_data_gather_conf.json: added spreadsheet settings")
        self.add_change(
            '2020-12-14', 'covid10_data_gather_conf.json: global settings now controlled')
        self.add_change(
            '2020-12-11', 'Fixed trend bug of always being in the DANGER ZONE; % growth or daily avg controls trend description.')
        self.add_change(
            '2020-12-10', 'Added comma formatting to Today tab, Daily Avg col.')
        self.add_change(
            '2020-12-08', 'Updated trend col - the ratio of newly-sick to just-recovered is used to categorize contagion growth.')
        self.add_change(
            '2020-10-10', 'Added daily avg - the average increase/decrease for the last seven days of cases for a geography.')
        self.add_change('2020-10-07', 'Split Data by County and State')
        self.add_change(
            '2020-07-16', 'Ignore Counties without parent geography (eg PR)')
        self.add_change('2020-07-09', 'Updated Per Capita => Per 100k.')
        self.add_change(
            '2020-07-08', 'Fixed sorting bug on today tab and added active per capita, tweaked number formats. ')
        self.add_change(
            '2020-07-07', 'Optimized active to inactive date mapping via cache.')
        self.add_change(
            '2020-07-06', 'Added active case count based on 28 day infection period. Day 0 now represents 1 case.')
        self.add_change('2020-05-02', 'County per capita scale is now 100k, matching states; each state now has an Unknown County to capture gaps in state data. PR/Gua/VI etc are still excluded.')
        self.add_change(
            '2020-04-28', 'Sort summary tab by new cases column instead of dead per capita')
        self.add_change(
            '2020-04-26', 'Added daily reported/dead delta & change log worksheets')
        self.add_change(
            '2020-04-25', 'Recalibrated minimum benchmark to 1000 from 10; Day "0" now represents 1000 reported cases, which will filter less infected geographies.')
        self.add_change(
            '2020-04-19', 'Reconfirmed per capita scaling of 10k for county, 100k for state; true per capita is not meaningful at current case rates.')
        self.add_change(
            '2020-04-18', 'Added Case Fatality Rate = Dead / Actual %. Refactored NYT data automation.')
        self.add_change(
            '2020-04-17', 'Added summary tab, removed redudant population tab.')
        self.add_change(
            '2020-04-10', 'Removed unreported infection tab as it was too speculative.')
        self.add_change(
            '2020-04-05', 'Refactored static geography set into objects--now produce multiple spreadsheets')
        self.add_change(
            '2020-03-29', 'Initial cut with simple/static geography set; day "0" is 10 cases. Assume unreported rate of 85%.')

    def __init__(self, locations, benchmark=g_case_benchmark, filename="data"):
        self.locations = locations
        self.benchmark = benchmark

        self.init_data()

        self.init_wb()
        self.init_ws()
        self.add_headers()

        self.gen_locations_data()
        self.add_counts_to_xlsx()

        self.add_population_to_xlsx()
        self.add_changelog_to_xlsx()

        # xslx filename handling
        xlsx_f = "covid19_%s_%s_data.xlsx" % (
            date.today().strftime("%Y_%m_%d"), filename)
        xlsx_fn = g_xlsx_path / xlsx_f
        self.wb_file = str(xlsx_fn.resolve())

        self.wb.save(self.wb_file)


def get_county_id(s, c, str):
    csv = str.split(",")
    county = csv[0].strip()
    state_abbr = csv[1].strip()
    state_fips = s.states_by_name[g_state_name[state_abbr]].state_fips
    #print("%s (%i) %s" % (state_abbr, state_fips, county))
    return c.counties_for_pop_est[state_fips][county]


def send_email_win_outlook(xlsx_files):
    olMailItem = 0x0
    today = date.today().strftime("%m/%d")
    obj = win32com.client.Dispatch("Outlook.Application")
    daily_email = obj.CreateItem(olMailItem)
    daily_email.Subject = "%s Covid19 Data" % (today)
    # olFormatHTML https://msdn.microsoft.com/en-us/library/office/aa219371(v=office.11).aspx
    daily_email.BodyFormat = 2
    daily_email.HTMLBody = """
    <HTML>
        <HEAD>
            <meta http-equiv="Content-type" content="text/html; charset=utf-8" />
        </HEAD>
        <BODY>
            <div style="%s">
                %s
                <p>
                Attached are today’s %s Covid19 stats pulled from the NYT!<p>

                %s
            </div>
        </BODY>
    </HTML>
    """ % (g_email_style, g_email_greeting, today, g_email_sig)
    if len(g_email_to) > 0:
        daily_email.To = ",".join(g_email_to)

    for xlsx in xlsx_files:
        xlsx_filename = xlsx.wb_file
        dbg(xlsx_filename)
        daily_email.Attachments.Add(Source=xlsx_filename)
    daily_email.Display(True)


def send_email(xlsx_files):
    # Only send if email flag is True:
    if not g_email:
        print(" -- no email sent per conf.json ")
        return

    # Todo, enum?
    # todo: unix/osx Mail/win Mail
    if g_email_client == "Outlook":
        send_email_win_outlook(xlsx_files)
    else:
        print(" -- no email sent, unrecognized email client: %s " %
              (g_email_client))


if __name__ == "__main__":

    br()
    t_main = log_start('Start daily covid19 processing.')

    br()
    load_configuration()
    update_data()
    cache_active_to_inactive_date_map()

    br()
    t = log_start("Load geography")
    s = States()
    c = Counties(s)
    set_county_population(s, c)
    validate_custom_geographies(s, c)
    log_end(t)

    br()
    t = log_start("Process covid data")
    set_county_covid19_cases(s, c)
    log_end(t)

    br()
    t = log_start("Extract sheet data")

    state_data = dict()
    custom_data = dict()

    for state_abbr in g_conf['spreadsheets']['state-detail']:
        # todo - validate state_abbr
        state_data[state_abbr] = s.get_by_name(
            g_state_name[state_abbr]).get_all_counties()

    for custom_xlsx in g_conf['spreadsheets']['custom']:
        custom_data[custom_xlsx] = []
        for geography in g_conf['spreadsheets']['custom'][custom_xlsx]:
            if len(geography) == 2:
                custom_data[custom_xlsx].append(
                    s.get_by_name(g_state_name[geography]))
            else:
                custom_data[custom_xlsx].append(get_county_id(s, c, geography))

    log_end(t)

    br()

    t = log_start("Generate xlsx")
    xlsx_files = []

    for xlsx_fn in custom_data.keys():
        xlsx_files.append(XLSX(custom_data[xlsx_fn], filename=xlsx_fn))

    # Generate US xlsx?
    if get_global_conf('spreadsheets', 'us', False):
        all_states = s.get_all_states()
        xlsx_files.append(XLSX(all_states, filename="US"))

    # Generate state XLSX?
    for state_abbr in state_data.keys():
        xlsx_files.append(XLSX(state_data[state_abbr], filename=state_abbr))

    log_end(t)

    br()
    t = log_start("Send email")
    send_email(xlsx_files)
    log_end(t)

    br()
    log_start('All done!')
    log_end(t_main)
